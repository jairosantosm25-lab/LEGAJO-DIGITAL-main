# app/application/services/legajo_service.py
# Importa la librería para calcular hashes de archivos.
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io
from flask import current_app
from datetime import datetime, timedelta


# Define el servicio que contiene la lógica de negocio para los legajos.
class LegajoService:
    # El constructor inyecta las dependencias del repositorio de personal y el servicio de auditoría.
    def __init__(self, personal_repository, audit_service):
        self._personal_repo = personal_repository
        self._audit_service = audit_service

    # --- MÉTODOS DE CONSULTA (GETTERS) ---

    def get_tipos_documento_by_seccion(self, seccion_id):
        """Orquesta la obtención de tipos de documento filtrados por sección."""
        return self._personal_repo.get_tipos_documento_by_seccion(seccion_id)

    def get_document_for_download(self, document_id):
        """Recupera un documento de la base de datos y lo prepara para la descarga."""
        document_row = self._personal_repo.find_document_by_id(document_id)
        if not document_row:
            return None
        # El SP devuelve una fila con (nombre_archivo, archivo_binario)
        return {"filename": document_row[0], "data": document_row[1]}

    def check_if_dni_exists(self, dni):
        """Orquesta la verificación de la existencia de un DNI."""
        return self._personal_repo.check_dni_exists(dni)

    def get_all_personal_paginated(self, page, per_page, filters=None):
        """Obtiene una lista paginada y filtrada de personal."""
        return self._personal_repo.get_all_paginated(page, per_page, filters)

    def get_personal_details(self, personal_id):
        """Obtiene todos los detalles del legajo de una persona por su ID."""
        return self._personal_repo.get_full_legajo_by_id(personal_id)

    def get_documents_by_personal_id(self, personal_id):
        """Obtiene los documentos de un empleado."""
        return self._personal_repo.find_documents_by_personal_id(personal_id)

    # --- MÉTODOS PARA POBLAR FORMULARIOS ---

    def get_unidades_for_select(self):
        return self._personal_repo.get_unidades_for_select()

    def get_secciones_for_select(self):
        return self._personal_repo.get_secciones_for_select()

    def get_tipos_documento_for_select(self):
        return self._personal_repo.get_tipos_documento_for_select()

    # --- MÉTODOS DE OPERACIONES (CUD) ---

    def register_new_personal(self, form_data, creating_user_id):
        """Registra un nuevo empleado y audita la acción."""
        new_personal_id = self._personal_repo.create(form_data)
        self._audit_service.log(creating_user_id, 'Personal', 'CREAR', f"Se creó el legajo para el DNI {form_data['dni']}", form_data)
        return new_personal_id

    def upload_document_to_personal(self, form_data, file_storage, current_user_id):
        """Gestiona la validación y subida de un nuevo documento."""
        if not file_storage or not file_storage.filename:
            raise ValueError("No se proporcionó ningún archivo para subir.")

        filename = file_storage.filename
        
        allowed_extensions = current_app.config['ALLOWED_EXTENSIONS']
        if '.' not in filename or filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            raise ValueError(f"Tipo de archivo no permitido. Solo se aceptan: {', '.join(allowed_extensions)}")

        file_bytes = file_storage.read()
        if len(file_bytes) > current_app.config['MAX_CONTENT_LENGTH']:
            max_size_mb = current_app.config['MAX_CONTENT_LENGTH'] / (1024 * 1024)
            raise ValueError(f"El archivo es demasiado grande. El tamaño máximo es de {max_size_mb:.0f} MB.")
        
        file_storage.seek(0)
        
        file_hash = hashlib.sha256(file_bytes).hexdigest()
        
        doc_data = form_data.copy()
        doc_data['nombre_archivo'] = filename
        doc_data['hash_archivo'] = file_hash
        id_personal = doc_data.get('id_personal')

        self._personal_repo.add_document(doc_data, file_bytes)
        
        self._audit_service.log(
            current_user_id,
            'Documentos',
            'SUBIR',
            f"Subió el archivo '{filename}' al legajo del personal ID {id_personal}"
        )

    def delete_personal_by_id(self, personal_id, deleting_user_id):
        """Desactiva un legajo de personal y audita la acción."""
        persona = self._personal_repo.find_by_id(personal_id)
        if not persona:
            raise ValueError("La persona que intenta eliminar no existe.")

        self._personal_repo.delete_by_id(personal_id)
        self._audit_service.log(
            deleting_user_id,
            'Personal',
            'ELIMINAR (Desactivar)',
            f"Se desactivó el legajo del personal con DNI {persona['dni']}"
        )

    def delete_document_by_id(self, document_id, deleting_user_id):
        """Orquesta la eliminación lógica de un documento y lo audita."""
        self._personal_repo.delete_document_by_id(document_id)
        self._audit_service.log(
            deleting_user_id,
            'Documentos',
            'ELIMINAR (Lógico)',
            f"Se marcó como eliminado el documento con ID {document_id}"
        )

    # --- MÉTODOS DE REPORTES Y ESTADO ---
    
    def generate_general_report_excel(self):
        """Genera un reporte general de personal en un archivo Excel."""
        personal_data = self._personal_repo.get_all_for_report()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte General de Personal"

        headers = [
            "DNI", "Apellidos", "Nombres", "Sexo", "Fecha de Nacimiento", "Email",
            "Teléfono", "Unidad Administrativa", "Fecha de Ingreso", "Estado",
            "Último Cargo", "Último Tipo de Contrato", "Modalidad", "Sueldo", "Resolución"
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for persona in personal_data:
            row_data = [
                persona.get('dni'), persona.get('apellidos'), persona.get('nombres'),
                persona.get('sexo'), persona.get('fecha_nacimiento'), persona.get('email'),
                persona.get('telefono'), persona.get('nombre_unidad'),
                persona.get('fecha_ingreso'), 'Activo' if persona.get('activo') else 'Inactivo',
                persona.get('cargo'), persona.get('tipo_contrato'), persona.get('modalidad'),
                persona.get('sueldo'), persona.get('resolucion')
            ]
            ws.append(row_data)

        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        
        return excel_stream

    def check_document_status_for_all_personal(self, days_to_expire=30):
        """Revisa documentos con fecha de vencimiento y resume el estado por persona."""
        all_docs = self._personal_repo.get_all_documents_with_expiration()
        status_summary = {}
        today = datetime.now().date()
        expiration_threshold = today + timedelta(days=days_to_expire)

        for doc in all_docs:
            personal_id = doc['id_personal']
            vencimiento = doc['fecha_vencimiento']

            if personal_id not in status_summary:
                status_summary[personal_id] = {'expired': 0, 'expiring_soon': 0}

            if vencimiento < today:
                status_summary[personal_id]['expired'] += 1
            elif vencimiento <= expiration_threshold:
                status_summary[personal_id]['expiring_soon'] += 1
        
        return status_summary